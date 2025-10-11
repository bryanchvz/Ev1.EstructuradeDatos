from datetime import datetime
import json
import csv
import os

salas = {}
reservaciones = {}
clientes = {}
contador_clientes = 0

ARCHIVO_DATOS = "datos_sistema.json"

def cargar_estado():
    global salas, reservaciones, clientes, contador_clientes
    
    if os.path.exists(ARCHIVO_DATOS):
        try:
            with open(ARCHIVO_DATOS, 'r', encoding='utf-8') as archivo:
                datos = json.load(archivo)
                salas = datos.get('salas', {})
                reservaciones = datos.get('reservaciones', {})
                clientes = datos.get('clientes', {})
                contador_clientes = datos.get('contador_clientes', 0)
            print("Estado anterior cargado exitosamente.")
        except:
            print("No se pudo cargar el estado anterior. Iniciando con estado vacío.")
    else:
        print("No se encontró un estado anterior. Iniciando con estado vacío.")

def guardar_estado():
    try:
        datos = {
            'salas': salas,
            'reservaciones': reservaciones,
            'clientes': clientes,
            'contador_clientes': contador_clientes
        }
        with open(ARCHIVO_DATOS, 'w', encoding='utf-8') as archivo:
            json.dump(datos, archivo, ensure_ascii=False, indent=4)
        print("Estado guardado exitosamente.")
        return True
    except Exception as e:
        print(f"Error al guardar el estado: {e}")
        return False

def registrar_cliente():
    global contador_clientes
    
    while True:
        nombre = input("Ingrese su nombre: ").strip()
        if nombre and nombre.replace(" ", "").isalpha():
            break
        print("Error: El nombre solo puede contener letras y no puede estar vacío.")

    while True:
        apellidos = input("Ingrese sus apellidos: ").strip()
        if apellidos and apellidos.replace(" ", "").isalpha():
            break
        print("Error: Los apellidos solo pueden contener letras y no pueden estar vacíos.")

    contador_clientes += 1
    clave = f"C{contador_clientes:03d}"
    clientes[clave] = {"nombre": nombre, "apellidos": apellidos}

    print("Registrado exitosamente.")

def listar_clientes():
    if not clientes:
        print("No hay clientes registrados.")
        return
    
    print("\n" + "="*100)
    print("LISTA DE CLIENTES REGISTRADOS")
    print("="*100)
    
    clientes_ordenados = sorted(clientes.items(), 
                               key=lambda x: (x[1]["apellidos"].lower(), x[1]["nombre"].lower()))
    
    for clave, datos in clientes_ordenados:
        print(f"{clave}: {datos['apellidos']}, {datos['nombre']}")
    print("="*100)

def registrar_sala():
    clave = f"S{len(salas)+1:03d}"
    
    while True:
        nombre = input("Ingrese el nombre de la sala: ").strip()
        if nombre:
            break
        print("Error: El nombre de la sala no puede estar vacío.")
    
    while True:
        try:
            cupo = int(input("Ingrese el cupo máximo de la sala: "))
            if cupo > 0:
                break
            print("Error: El cupo debe ser mayor a 0.")
        except ValueError:
            print("Error: Debe ingresar un número válido.")
    
    salas[clave] = {"nombre": nombre, "cupo": cupo}
    print("Registrado exitosamente.")

def mostrar_salas_disponibles(fecha_str):
    if not salas:
        return {}
    
    salas_disponibles = {}
    print(f"\nSalas disponibles para {fecha_str}:")
    print("-" * 100)
    
    for clave_sala, info_sala in salas.items():
        turnos_disponibles = ["MATUTINO", "VESPERTINO", "NOCTURNO"]
        
        for reserva in reservaciones.values():
            if reserva["sala"] == clave_sala and reserva["fecha"] == fecha_str:
                if reserva["turno"] in turnos_disponibles:
                    turnos_disponibles.remove(reserva["turno"])
        
        if turnos_disponibles:
            salas_disponibles[clave_sala] = turnos_disponibles
            print(f"{clave_sala}: {info_sala['nombre']} (cupo {info_sala['cupo']}) - Turnos: {', '.join(turnos_disponibles)}")
    
    if not salas_disponibles:
        print("No hay salas disponibles para esta fecha.")
    
    print("-" * 100)
    return salas_disponibles

def registrar_reservacion():
    if not clientes:
        print("Error: Primero debe registrar clientes.")
        return
    if not salas:
        print("Error: Primero debe registrar salas.")
        return

    listar_clientes()
    cliente = None
    while not cliente:
        clave_input = input("\nIngrese la clave del cliente: ").strip().upper()
        if clave_input in clientes:
            cliente = clave_input
        else:
            print("Error: Clave de cliente inválida.")
            listar_clientes()

    fecha_str = None
    while not fecha_str:
        fecha_input = input("Ingrese la fecha de reserva (DD-MM-YYYY): ").strip()
        try:
            fecha = datetime.strptime(fecha_input, "%d-%m-%Y")
            hoy = datetime.now()
            diferencia_dias = (fecha - hoy).days
            if diferencia_dias >= 2:
                fecha_str = fecha_input
            else:
                print("Error: La fecha debe ser al menos dos días después de hoy.")
        except ValueError:
            print("Error: Formato de fecha inválido. Use DD-MM-YYYY")

    salas_disponibles = mostrar_salas_disponibles(fecha_str)
    if not salas_disponibles:
        return

    sala = None
    while not sala:
        sala_input = input("Ingrese la clave de la sala: ").strip().upper()
        if sala_input in salas_disponibles:
            sala = sala_input
        else:
            print("Error: Clave de sala inválida o no disponible.")

    turnos_validos = salas_disponibles[sala]
    print(f"Turnos disponibles: {', '.join(turnos_validos)}")
    turno = None
    while not turno:
        turno_input = input("Ingrese el turno (MATUTINO/VESPERTINO/NOCTURNO): ").strip().upper()
        if turno_input in turnos_validos:
            turno = turno_input
        else:
            print("Error: Turno no válido o no disponible.")

    evento = None
    while not evento:
        evento_input = input("Ingrese el nombre del evento: ").strip()
        if evento_input:
            evento = evento_input
        else:
            print("Error: El nombre del evento no puede estar vacío.")

    folio = f"R{len(reservaciones)+1:03d}"
    reservaciones[folio] = {
        "cliente": cliente,
        "sala": sala,
        "fecha": fecha_str,
        "turno": turno,
        "evento": evento
    }
    
    print("Registrado exitosamente.")

def editar_evento():
    if not reservaciones:
        print("No hay reservaciones registradas.")
        return
    
    while True:
        fecha_inicio = input("Ingrese la fecha inicial del rango (DD-MM-YYYY): ").strip()
        try:
            fecha_inicio_obj = datetime.strptime(fecha_inicio, "%d-%m-%Y")
            break
        except ValueError:
            print("Error: Formato de fecha inválido. Use DD-MM-YYYY")
    
    while True:
        fecha_fin = input("Ingrese la fecha final del rango (DD-MM-YYYY): ").strip()
        try:
            fecha_fin_obj = datetime.strptime(fecha_fin, "%d-%m-%Y")
            if fecha_fin_obj >= fecha_inicio_obj:
                break
            print("Error: La fecha final debe ser mayor o igual a la fecha inicial.")
        except ValueError:
            print("Error: Formato de fecha inválido. Use DD-MM-YYYY")
    
    reservas_en_rango = []
    for folio, reserva in reservaciones.items():
        try:
            fecha_reserva = datetime.strptime(reserva["fecha"], "%d-%m-%Y")
            if fecha_inicio_obj <= fecha_reserva <= fecha_fin_obj:
                reservas_en_rango.append((folio, reserva))
        except:
            pass
    
    if not reservas_en_rango:
        print(f"No hay reservaciones en el rango {fecha_inicio} a {fecha_fin}")
        return
    
    print("\n" + "="*140)
    print(f"RESERVACIONES DEL {fecha_inicio} AL {fecha_fin}")
    print("="*140)
    print(f"{'FOLIO':<10} {'EVENTO':<35} {'CLIENTE':<30} {'SALA':<30} {'FECHA':<15} {'TURNO':<15}")
    print("-"*140)
    
    for folio, reserva in reservas_en_rango:
        cliente_info = clientes[reserva["cliente"]]
        sala_info = salas[reserva["sala"]]
        cliente_nombre = f"{cliente_info['nombre']} {cliente_info['apellidos']}"
        
        evento_truncado = reserva['evento'][:33] + '..' if len(reserva['evento']) > 35 else reserva['evento']
        cliente_truncado = cliente_nombre[:28] + '..' if len(cliente_nombre) > 30 else cliente_nombre
        sala_truncada = sala_info['nombre'][:28] + '..' if len(sala_info['nombre']) > 30 else sala_info['nombre']
        
        print(f"{folio:<10} {evento_truncado:<35} {cliente_truncado:<30} {sala_truncada:<30} {reserva['fecha']:<15} {reserva['turno']:<15}")
    
    print("="*140)
    
    folio = None
    while not folio:
        folio_input = input("\nIngrese el folio de la reservación a editar: ").strip().upper()
        if folio_input in reservaciones:
            folio = folio_input
        else:
            print("Error: Folio de reservación inválido.")
    
    print(f"Evento actual: {reservaciones[folio]['evento']}")
    nuevo_evento = None
    while not nuevo_evento:
        nuevo_input = input("Ingrese el nuevo nombre del evento: ").strip()
        if nuevo_input:
            reservaciones[folio]['evento'] = nuevo_input
            print("Registrado exitosamente.")
            nuevo_evento = nuevo_input
        else:
            print("Error: El nombre del evento no puede estar vacío.")

def exportar_csv(fecha_reporte, reservas_fecha):
    nombre_archivo = f"reporte_{fecha_reporte.replace('-', '_')}.csv"
    
    try:
        with open(nombre_archivo, 'w', newline='', encoding='utf-8') as archivo:
            escritor = csv.writer(archivo, delimiter= "|")
            escritor.writerow(['SALA', 'CLIENTE', 'EVENTO', 'TURNO'])
            
            reservas_ordenadas = sorted(reservas_fecha, key=lambda x: int(x[1]["sala"][1:]))
            
            for folio, reserva in reservas_ordenadas:
                sala_num = reserva["sala"][1:]
                cliente_info = clientes[reserva["cliente"]]
                cliente_nombre = f"{cliente_info['apellidos']} {cliente_info['nombre']}"
                escritor.writerow([sala_num, cliente_nombre, reserva['evento'], reserva['turno']])
        
        print(f"Reporte CSV exportado exitosamente: {nombre_archivo}")
    except Exception as e:
        print(f"Error al exportar CSV: {e}")

def exportar_json(fecha_reporte, reservas_fecha):
    nombre_archivo = f"reporte_{fecha_reporte.replace('-', '_')}.json"
    
    try:
        reporte = {
            "fecha": fecha_reporte,
            "reservaciones": []
        }
        
        reservas_ordenadas = sorted(reservas_fecha, key=lambda x: int(x[1]["sala"][1:]))
        
        for folio, reserva in reservas_ordenadas:
            sala_num = reserva["sala"][1:]
            cliente_info = clientes[reserva["cliente"]]
            cliente_nombre = f"{cliente_info['apellidos']} {cliente_info['nombre']}"
            
            reporte["reservaciones"].append({
                "sala": sala_num,
                "cliente": cliente_nombre,
                "evento": reserva['evento'],
                "turno": reserva['turno']
            })
        
        with open(nombre_archivo, 'w', encoding='utf-8') as archivo:
            json.dump(reporte, archivo, ensure_ascii=False, indent=4)
        
        print(f"Reporte JSON exportado exitosamente: {nombre_archivo}")
    except Exception as e:
        print(f"Error al exportar JSON: {e}")

def exportar_excel(fecha_reporte, reservas_fecha):
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side
    except ImportError:
        print("Error: La biblioteca openpyxl no está instalada.")
        print("Instale con: pip install openpyxl")
        return
    
    nombre_archivo = f"reporte_{fecha_reporte.replace('-', '_')}.xlsx"
    
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporte de Reservaciones"
        
        ws['A1'] = f"REPORTE DE RESERVACIONES PARA EL DÍA {fecha_reporte}"
        ws.merge_cells('A1:D1')
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        encabezados = ['SALA', 'CLIENTE', 'EVENTO', 'TURNO']
        for col, encabezado in enumerate(encabezados, start=1):
            celda = ws.cell(row=3, column=col)
            celda.value = encabezado
            celda.font = Font(bold=True)
            celda.alignment = Alignment(horizontal='center', vertical='center')
        
        reservas_ordenadas = sorted(reservas_fecha, key=lambda x: int(x[1]["sala"][1:]))
        
        borde = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        fila = 4
        for folio, reserva in reservas_ordenadas:
            sala_num = reserva["sala"][1:]
            cliente_info = clientes[reserva["cliente"]]
            cliente_nombre = f"{cliente_info['apellidos']} {cliente_info['nombre']}"
            
            ws.cell(row=fila, column=1, value=sala_num).alignment = Alignment(horizontal='center')
            ws.cell(row=fila, column=2, value=cliente_nombre).alignment = Alignment(horizontal='center')
            ws.cell(row=fila, column=3, value=reserva['evento']).alignment = Alignment(horizontal='center')
            ws.cell(row=fila, column=4, value=reserva['turno']).alignment = Alignment(horizontal='center')
            
            for col in range(1, 5):
                ws.cell(row=fila, column=col).border = borde
            
            fila += 1
        
        for col in range(1, 5):
            ws.cell(row=3, column=col).border = borde
        
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 15
        
        wb.save(nombre_archivo)
        print(f"Reporte Excel exportado exitosamente: {nombre_archivo}")
    except Exception as e:
        print(f"Error al exportar Excel: {e}")

def consultar_reservaciones():
    if not reservaciones:
        print("No hay reservaciones registradas.")
        return
    
    while True:
        fecha_consulta = input("Ingrese la fecha a consultar (DD-MM-YYYY): ").strip()
        try:
            datetime.strptime(fecha_consulta, "%d-%m-%Y")
            break
        except ValueError:
            print("Error: Formato de fecha inválido. Use DD-MM-YYYY")
    
    reservas_fecha = []
    for folio, reserva in reservaciones.items():
        if reserva["fecha"] == fecha_consulta:
            reservas_fecha.append((folio, reserva))
    
    if not reservas_fecha:
        print(f"No hay reservaciones para la fecha {fecha_consulta}")
        return
    
    print(f"\n" + "*" * 120)
    print("**" + " " * 116 + "**")
    print(f"**{f'REPORTE DE RESERVACIONES PARA EL DÍA {fecha_consulta}':^116}**")
    print("**" + " " * 116 + "**")
    print("*" * 120)
    print(f"{'SALA':<10} {'CLIENTE':<35} {'EVENTO':<45} {'TURNO'}")
    print("*" * 120)
    
    reservas_ordenadas = sorted(reservas_fecha, key=lambda x: int(x[1]["sala"][1:]))
    
    for folio, reserva in reservas_ordenadas:
        sala_num = reserva["sala"][1:]
        cliente_info = clientes[reserva["cliente"]]
        cliente_nombre = f"{cliente_info['apellidos']} {cliente_info['nombre']}"
        
        print(f"{sala_num:<10} {cliente_nombre:<35} {reserva['evento']:<45} {reserva['turno']}")
    
    print("*" * 51 + " FIN DEL REPORTE " + "*" * 51)
    print()
    
    print("\n¿Desea exportar este reporte?")
    print("1. CSV")
    print("2. JSON")
    print("3. Excel")
    print("4. No exportar")
    
    opcion_exportar = input("Seleccione una opción (1-4): ").strip()
    
    if opcion_exportar == "1":
        exportar_csv(fecha_consulta, reservas_fecha)
    elif opcion_exportar == "2":
        exportar_json(fecha_consulta, reservas_fecha)
    elif opcion_exportar == "3":
        exportar_excel(fecha_consulta, reservas_fecha)
    elif opcion_exportar == "4":
        print("No se exportó el reporte.")
    else:
        print("Opción inválida. No se exportó el reporte.")

def mostrar_menu():
    print("\n" + "="*60)
    print("SISTEMA DE RESERVAS DE SALAS DE COWORKING")
    print("="*60)
    print("1. Registrar reservación de una sala")
    print("2. Editar el nombre del evento de una reservación")
    print("3. Consultar reservaciones por fecha")
    print("4. Registrar un nuevo cliente")
    print("5. Registrar una sala")
    print("6. Salir")
    print("="*60)

def main():
    print("¡Bienvenido al Sistema de Reservas de Salas de Coworking!")
    print("\nCargando estado anterior...")
    cargar_estado()
    
    while True:
        mostrar_menu()
        
        opcion = input("Seleccione una opción (1-6): ").strip()
        
        if opcion == "1":
            registrar_reservacion()
        elif opcion == "2":
            editar_evento()
        elif opcion == "3":
            consultar_reservaciones()
        elif opcion == "4":
            registrar_cliente()
        elif opcion == "5":
            registrar_sala()
        elif opcion == "6":
            confirmacion = input("\n¿Está seguro que desea salir? (s/n): ").strip().lower()
            if confirmacion in ['s', 'si', 'sí', 'y', 'yes']:
                print("\nGuardando estado de la solución...")
                if guardar_estado():
                    print("¡Gracias por usar el Sistema de Reservas!")
                    print("Saliendo del programa...")
                    break
                else:
                    print("Hubo un error al guardar. ¿Desea salir de todos modos? (s/n): ")
                    salir_forzado = input().strip().lower()
                    if salir_forzado in ['s', 'si', 'sí', 'y', 'yes']:
                        print("Saliendo sin guardar...")
                        break
            else:
                print("Regresando al menú principal...")
        else:
            print("Error: Opción inválida. Por favor seleccione una opción del 1 al 6.")

if __name__ == "__main__":
    main()